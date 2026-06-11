from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .normalization import clean_text, normalize_header


MARKER_ALIASES = {
    "supplier": {"采购类型", "供应商", "供应商名称"},
    "customer": {"采购负责人", "负责人", "客户", "学校"},
}

MARK_COLORS = (
    "FFFFFF00",  # yellow
    "FF92D050",  # green
    "FF00B0F0",  # blue
    "FFFFC000",  # orange
    "FFD9EAD3",  # light green
    "FFCFE2F3",  # light blue
    "FFF4CCCC",  # light red
    "FFEADCF8",  # light purple
    "FFFFE599",  # light yellow
    "FFD9D2E9",  # lavender
    "FFB6D7A8",  # muted green
    "FFA4C2F4",  # muted blue
)

WHITE_FILL = PatternFill(fill_type=None)


@dataclass(frozen=True)
class SupplierAnnotationSummary:
    total_rows: int
    matched_rows: int
    supplier_count: int
    customer_count: int
    suppliers: list[str]
    customers: list[str]


def annotate_supplier_purchase_order(
    input_path: str | Path,
    output_path: str | Path,
    *,
    supplier_keyword: str = "",
    keep_first_customer_uncolored: bool = True,
) -> SupplierAnnotationSummary:
    input_path = Path(input_path)
    output_path = Path(output_path)
    wb = load_workbook(input_path)
    ws = wb.active
    header_row, columns = _find_marker_header(ws)
    supplier_col = columns["supplier"]
    customer_col = columns["customer"]

    supplier_keyword = clean_text(supplier_keyword)
    matched_rows_by_supplier: dict[str, list[int]] = {}
    customers_by_supplier: dict[str, list[str]] = {}
    all_customers: list[str] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        supplier = _supplier_name(ws.cell(row_idx, supplier_col).value)
        customer = clean_text(ws.cell(row_idx, customer_col).value)
        if not supplier or not customer:
            continue
        if supplier_keyword and supplier_keyword not in supplier and supplier_keyword not in clean_text(ws.cell(row_idx, supplier_col).value):
            continue
        matched_rows_by_supplier.setdefault(supplier, []).append(row_idx)
        if customer not in customers_by_supplier.setdefault(supplier, []):
            customers_by_supplier[supplier].append(customer)
        if customer not in all_customers:
            all_customers.append(customer)

    for supplier, row_indices in matched_rows_by_supplier.items():
        customer_colors = _customer_color_map(
            customers_by_supplier[supplier],
            keep_first_customer_uncolored=keep_first_customer_uncolored,
        )
        for row_idx in row_indices:
            customer = clean_text(ws.cell(row_idx, customer_col).value)
            fill = customer_colors.get(customer)
            if fill is None:
                continue
            ws.cell(row_idx, customer_col).fill = fill

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    suppliers = list(matched_rows_by_supplier)
    return SupplierAnnotationSummary(
        total_rows=max(ws.max_row - header_row, 0),
        matched_rows=sum(len(rows) for rows in matched_rows_by_supplier.values()),
        supplier_count=len(suppliers),
        customer_count=len(all_customers),
        suppliers=suppliers,
        customers=all_customers,
    )


def _find_marker_header(ws: Worksheet) -> tuple[int, dict[str, int]]:
    normalized_aliases = {
        key: {normalize_header(alias) for alias in values}
        for key, values in MARKER_ALIASES.items()
    }
    best_row = 0
    best_columns: dict[str, int] = {}
    required = set(MARKER_ALIASES)

    for row_number in range(1, min(ws.max_row, 30) + 1):
        found: dict[str, int] = {}
        for col_number in range(1, ws.max_column + 1):
            normalized = normalize_header(ws.cell(row_number, col_number).value)
            if not normalized:
                continue
            for key, possible in normalized_aliases.items():
                if normalized in possible and key not in found:
                    found[key] = col_number
        if len(required & found.keys()) > len(required & best_columns.keys()):
            best_row = row_number
            best_columns = found
        if required.issubset(found.keys()):
            return row_number, found

    missing = ", ".join(sorted(required - best_columns.keys()))
    raise ValueError(f"找不到供应商采购单标注所需表头：{missing}")


def _supplier_name(value) -> str:
    text = clean_text(value)
    if "/" in text:
        return clean_text(text.split("/", 1)[1])
    return text


def _customer_color_map(customers: list[str], *, keep_first_customer_uncolored: bool) -> dict[str, PatternFill]:
    result: dict[str, PatternFill] = {}
    color_index = 0
    for index, customer in enumerate(customers):
        if keep_first_customer_uncolored and index == 0:
            result[customer] = WHITE_FILL
            continue
        color = MARK_COLORS[color_index % len(MARK_COLORS)]
        result[customer] = PatternFill(fill_type="solid", fgColor=color)
        color_index += 1
    return result
