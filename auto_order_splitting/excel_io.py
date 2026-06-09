from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    OrderLine,
    PurchaseLine,
    SkippedLine,
    TemplateItem,
    UnmatchedLine,
)
from .normalization import clean_text, decimal_to_excel, normalize_header, to_decimal


TEMPLATE_ALIASES = {
    "owner": {"负责人"},
    "supplier": {"采购员/供应商名称", "供应商名称"},
    "product": {"商品名称"},
    "unit": {"商品单位", "单位"},
    "quantity": {"采购数量", "计划采购量"},
    "price": {"采购单价", "单价"},
    "note": {"商品备注", "备注"},
    "order_note": {"订单备注"},
    "agreement_price": {"已设置采购协议价", "采购协议价"},
}

ORDER_ALIASES = {
    "order_no": {"订单号"},
    "customer": {"客户名称", "客户名称(学校)", "学校"},
    "ship_date": {"发货日期", "日期"},
    "product": {"商品名称"},
    "quantity": {"发货数量", "下单数量", "订购数量", "订单数量", "数量"},
    "amount": {"实际金额", "下单金额", "订单金额"},
    "subtotal": {"发货小计", "小计"},
    "unit": {"发货单位", "下单单位", "订购单位", "订单单位", "单位"},
    "order_price": {"发货单价", "下单单价", "订购单价", "订单单价", "单价"},
    "product_note": {"商品备注"},
    "order_note": {"订单备注", "备注"},
    "supplier": {"默认供应商", "供应商", "供应商名称"},
}

NOTE_LABEL_PREFIXES = ("商品备注：", "商品备注:", "订单备注：", "订单备注:")


def read_template(path: str | Path) -> list[TemplateItem]:
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row, columns = _find_header(ws, TEMPLATE_ALIASES, required={"supplier", "product", "price"})
    items: list[TemplateItem] = []

    for row_number in range(header_row + 1, ws.max_row + 1):
        row = _row_dict(ws, row_number, columns)
        product = clean_text(row.get("product"))
        supplier = clean_text(row.get("supplier"))
        if not product and not supplier:
            continue
        if not product or not supplier:
            continue
        items.append(
            TemplateItem(
                row_number=row_number,
                owner=clean_text(row.get("owner")),
                supplier=supplier,
                product=product,
                unit=clean_text(row.get("unit")),
                quantity=to_decimal(row.get("quantity")),
                price=row.get("price"),
                note=clean_text(row.get("note")),
                agreement_price=row.get("agreement_price"),
            )
        )
    return items


def read_orders(paths: Iterable[str | Path]) -> list[OrderLine]:
    lines: list[OrderLine] = []
    for raw_path in paths:
        path = Path(raw_path)
        wb = load_workbook(path, data_only=True)
        ws, header_row, columns = _find_header_in_workbook(
            wb,
            ORDER_ALIASES,
            required={"order_no", "customer", "product", "quantity", "unit"},
        )

        for row_number in range(header_row + 1, ws.max_row + 1):
            row = _row_dict(ws, row_number, columns)
            first_cell = clean_text(ws.cell(row_number, 1).value)
            if not first_cell and not clean_text(row.get("product")):
                continue
            if first_cell.startswith("小计"):
                continue
            product = clean_text(row.get("product"))
            if not product:
                continue
            product_note = clean_text(row.get("product_note"))
            order_note = clean_text(row.get("order_note"))
            lines.append(
                OrderLine(
                    row_number=row_number,
                    order_no=clean_text(row.get("order_no")),
                    customer=clean_text(row.get("customer")),
                    ship_date=row.get("ship_date"),
                    product=product,
                    quantity=to_decimal(row.get("quantity")),
                    amount=row.get("amount"),
                    subtotal=row.get("subtotal"),
                    unit=clean_text(row.get("unit")),
                    order_price=row.get("order_price"),
                    note=_order_output_note(product_note, order_note),
                    source_file=path,
                    supplier=clean_text(row.get("supplier")),
                    product_note=product_note,
                    order_note=order_note,
                )
            )
    return lines


def write_purchase_import(template_path: str | Path, output_path: str | Path, lines: list[PurchaseLine]) -> None:
    template_path = Path(template_path)
    output_path = Path(output_path)
    wb = load_workbook(template_path)
    ws = wb.active
    header_row, columns = _find_header(ws, TEMPLATE_ALIASES, required={"supplier", "product", "price"})
    output_max_col = max(ws.max_column, _template_output_max_column(ws, header_row, columns))
    template_data_row = header_row + 1
    existing_max_row = ws.max_row
    output_last_row = header_row + len(lines)

    for row_idx, line in enumerate(lines, start=header_row + 1):
        if row_idx > existing_max_row:
            _copy_row_format(ws, template_data_row, row_idx, output_max_col)
        values = _purchase_line_values(line)
        for key, col_idx in columns.items():
            if col_idx <= output_max_col and key in values:
                ws.cell(row_idx, col_idx, values[key])

    _delete_extra_data_rows(ws, output_last_row + 1, existing_max_row)
    _apply_output_filter(ws, header_row, len(lines), output_max_col)
    _reset_output_view(ws, header_row)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_debug_workbook(
    output_path: str | Path,
    lines: list[PurchaseLine],
    unmatched: list[UnmatchedLine],
    skipped: list[SkippedLine],
    warnings: list[str],
) -> None:
    output_path = Path(output_path)
    wb = Workbook()

    purchase_ws = wb.active
    purchase_ws.title = "采购结果"
    _write_rows(
        purchase_ws,
        ["供应商", "学校", "商品", "单位", "数量", "单价", "商品备注", "订单备注", "备注", "匹配方式"],
        [
            [
                line.supplier,
                line.owner,
                line.product,
                line.unit,
                decimal_to_excel(line.quantity),
                line.price,
                line.product_note,
                line.order_note,
                line.note,
                "、".join(sorted(line.match_methods)),
            ]
            for line in lines
        ],
    )

    unmatched_ws = wb.create_sheet("未匹配")
    _write_rows(
        unmatched_ws,
        ["文件", "行", "订单号", "学校", "商品", "数量", "单位", "商品备注", "订单备注", "备注", "原因"],
        [
            [
                item.source_file.name,
                item.row_number,
                item.order_no,
                item.customer,
                item.product,
                decimal_to_excel(item.quantity),
                item.unit,
                item.product_note,
                item.order_note,
                item.note,
                item.reason,
            ]
            for item in unmatched
        ],
    )

    warnings_ws = wb.create_sheet("警告")
    _write_rows(warnings_ws, ["警告"], [[warning] for warning in warnings])

    skipped_ws = wb.create_sheet("跳过")
    _write_rows(
        skipped_ws,
        ["文件", "行", "订单号", "学校", "商品", "数量", "单位", "商品备注", "订单备注", "备注", "原因"],
        [
            [
                item.source_file.name,
                item.row_number,
                item.order_no,
                item.customer,
                item.product,
                decimal_to_excel(item.quantity),
                item.unit,
                item.product_note,
                item.order_note,
                item.note,
                item.reason,
            ]
            for item in skipped
        ],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _find_header(
    ws: Worksheet,
    aliases: dict[str, set[str]],
    *,
    required: set[str],
) -> tuple[int, dict[str, int]]:
    normalized_aliases = {
        key: {normalize_header(alias) for alias in values}
        for key, values in aliases.items()
    }
    best_row = 0
    best_columns: dict[str, int] = {}

    for row_number in range(1, min(ws.max_row, 30) + 1):
        found: dict[str, int] = {}
        for col_number in range(1, ws.max_column + 1):
            normalized = normalize_header(ws.cell(row_number, col_number).value)
            if not normalized:
                continue
            for key, possible in normalized_aliases.items():
                if normalized in possible and key not in found:
                    found[key] = col_number
        if len(required & found.keys()) > len(required & best_columns.keys()):
            best_row = row_number
            best_columns = found
        if required.issubset(found.keys()):
            return row_number, found

    missing = ", ".join(sorted(required - best_columns.keys()))
    raise ValueError(f"在工作表「{ws.title}」找不到必要表头：{missing}")


def _find_header_in_workbook(
    wb,
    aliases: dict[str, set[str]],
    *,
    required: set[str],
) -> tuple[Worksheet, int, dict[str, int]]:
    worksheets = [wb.active] + [ws for ws in wb.worksheets if ws is not wb.active]
    errors: list[str] = []
    for ws in worksheets:
        try:
            header_row, columns = _find_header(ws, aliases, required=required)
            return ws, header_row, columns
        except ValueError as error:
            errors.append(str(error))

    required_labels = "、".join(sorted(required))
    raise ValueError(f"找不到订单明细表，必要字段：{required_labels}。已检查：{'；'.join(errors[:4])}")


def _order_output_note(product_note: str, order_note: str) -> str:
    return _merge_output_notes(product_note, order_note)


def _merge_output_notes(*values: str) -> str:
    parts: list[str] = []
    for value in values:
        for part in _clean_note_parts(value):
            if part not in parts:
                parts.append(part)
    return "；".join(parts)


def _clean_output_note(value: str) -> str:
    return _merge_output_notes(value)


def _clean_note_parts(value: str) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    parts: list[str] = []
    for raw_part in text.replace("；", ";").replace("\n", ";").replace("\r", ";").split(";"):
        part = _strip_note_label(raw_part)
        if part:
            parts.append(part)
    return parts


def _strip_note_label(value: str) -> str:
    text = clean_text(value)
    for prefix in NOTE_LABEL_PREFIXES:
        if text.startswith(prefix):
            return clean_text(text[len(prefix) :])
    return text


def _row_dict(ws: Worksheet, row_number: int, columns: dict[str, int]) -> dict[str, Any]:
    return {key: ws.cell(row_number, col_number).value for key, col_number in columns.items()}


def _copy_cell(src, dst) -> None:
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    if src.hyperlink:
        dst._hyperlink = copy(src.hyperlink)
    if src.comment:
        dst.comment = copy(src.comment)


def _copy_cell_format(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def _copy_row_format(ws: Worksheet, source_row: int, target_row: int, max_col: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col_idx in range(1, max_col + 1):
        _copy_cell_format(ws.cell(source_row, col_idx), ws.cell(target_row, col_idx))


def _delete_extra_data_rows(ws: Worksheet, start_row: int, end_row: int) -> None:
    if start_row <= end_row:
        ws.delete_rows(start_row, end_row - start_row + 1)


def _reset_output_view(ws: Worksheet, header_row: int) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]


def _template_output_max_column(ws: Worksheet, header_row: int, columns: dict[str, int]) -> int:
    header_cols = [
        cell.column
        for cell in ws[header_row]
        if clean_text(cell.value)
    ]
    return max(header_cols + list(columns.values()))


def _purchase_line_values(line: PurchaseLine) -> dict[str, Any]:
    return {
        "owner": line.owner,
        "supplier": line.supplier,
        "product": line.product,
        "unit": line.unit,
        "quantity": decimal_to_excel(line.quantity),
        "price": line.price,
        "note": _merge_output_notes(line.note, line.product_note, line.order_note),
        "order_note": _clean_output_note(line.order_note),
        "agreement_price": line.agreement_price,
    }


def _apply_output_filter(ws: Worksheet, header_row: int, data_rows: int, max_col: int) -> None:
    if data_rows:
        last_col = get_column_letter(max_col)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + data_rows}"


def _write_rows(ws: Worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="38342D")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_length = max(len(clean_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 46)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
