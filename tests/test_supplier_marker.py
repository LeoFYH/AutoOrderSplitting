from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from auto_order_splitting.supplier_marker import annotate_supplier_purchase_order


class SupplierMarkerTests(unittest.TestCase):
    def test_marks_customers_for_selected_supplier_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "marked.xlsx"
            _write_purchase_export(input_path)

            summary = annotate_supplier_purchase_order(
                input_path,
                output_path,
                supplier_keyword="供应商A",
                keep_first_customer_uncolored=True,
            )

            ws = load_workbook(output_path).active
            self.assertEqual(summary.matched_rows, 3)
            self.assertEqual(summary.supplier_count, 1)
            self.assertEqual(summary.customer_count, 2)
            self.assertIsNone(ws.cell(2, 2).fill.fill_type)
            self.assertEqual(ws.cell(3, 2).fill.fill_type, "solid")
            self.assertEqual(ws.cell(3, 2).fill.fgColor.rgb, ws.cell(4, 2).fill.fgColor.rgb)
            self.assertIsNone(ws.cell(5, 2).fill.fill_type)

    def test_marks_each_supplier_group_independently_when_keyword_is_blank(self):
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "source.xlsx"
            output_path = Path(tmp) / "marked.xlsx"
            _write_purchase_export(input_path)

            summary = annotate_supplier_purchase_order(
                input_path,
                output_path,
                supplier_keyword="",
                keep_first_customer_uncolored=True,
            )

            ws = load_workbook(output_path).active
            self.assertEqual(summary.matched_rows, 4)
            self.assertEqual(summary.supplier_count, 2)
            self.assertEqual(summary.customer_count, 3)
            self.assertIsNone(ws.cell(2, 2).fill.fill_type)
            self.assertEqual(ws.cell(3, 2).fill.fill_type, "solid")
            self.assertIsNone(ws.cell(5, 2).fill.fill_type)


def _write_purchase_export(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "采购单导出"
    ws.append(["采购类型", "采购负责人", "采购单号", "交货日期", "商品名称", "单位", "计划采购量", "备注"])
    ws.append(["供应商/供应商A", "客户甲", "CG1", "明天", "鸡肉", "斤", 1, None])
    ws.append(["供应商/供应商A", "客户乙", "CG1", "明天", "牛肉", "斤", 2, None])
    ws.append(["供应商/供应商A", "客户乙", "CG1", "明天", "猪肉", "斤", 3, None])
    ws.append(["供应商/供应商B", "客户丙", "CG2", "明天", "豆腐", "斤", 4, None])
    wb.save(path)


if __name__ == "__main__":
    unittest.main()
