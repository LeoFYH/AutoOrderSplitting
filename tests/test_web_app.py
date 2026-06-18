from __future__ import annotations

from io import BytesIO
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from auto_order_splitting import web_app
from auto_order_splitting import web_state
from auto_order_splitting.excel_io import read_template
from auto_order_splitting.web_app import parse_skip_keywords


class WebAppTests(unittest.TestCase):
    def test_empty_skip_keywords_stay_empty(self):
        self.assertEqual(parse_skip_keywords(""), [])
        self.assertEqual(parse_skip_keywords(None), [])

    def test_skip_keywords_parse_multiple_values(self):
        self.assertEqual(parse_skip_keywords("营养餐，早晚餐, 食材配送"), ["营养餐", "早晚餐", "食材配送"])

    def test_supplier_annotation_endpoint_returns_download(self):
        with tempfile.TemporaryDirectory() as tmp:
            old_runs_dir = web_app.RUNS_DIR
            web_app.RUNS_DIR = Path(tmp) / "runs"
            try:
                app = web_app.create_app()
                client = app.test_client()
                response = client.post(
                    "/api/supplier-annotate",
                    data={
                        "purchase": (_purchase_export_bytes(), "purchase.xlsx"),
                        "supplierKeyword": "供应商A",
                        "keepFirstCustomerUncolored": "true",
                    },
                    content_type="multipart/form-data",
                )
            finally:
                web_app.RUNS_DIR = old_runs_dir

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["summary"]["matchedRows"], 2)
        self.assertEqual(payload["summary"]["customerCount"], 2)
        self.assertIn("供应商采购单标注.xlsx", payload["downloads"]["annotated"])

    def test_process_falls_back_to_saved_template_when_template_upload_is_an_order_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            data_dir = Path(tmp) / "data"
            old_web_app_runs_dir = web_app.RUNS_DIR
            old_web_app_current_template = web_app.CURRENT_TEMPLATE
            old_state_data_dir = web_state.DATA_DIR
            old_state_runs_dir = web_state.RUNS_DIR
            old_state_current_template = web_state.CURRENT_TEMPLATE
            old_state_template_json = web_state.TEMPLATE_JSON
            try:
                web_app.RUNS_DIR = data_dir / "runs"
                web_app.CURRENT_TEMPLATE = data_dir / "current_template.xlsx"
                web_state.DATA_DIR = data_dir
                web_state.RUNS_DIR = data_dir / "runs"
                web_state.CURRENT_TEMPLATE = data_dir / "current_template.xlsx"
                web_state.TEMPLATE_JSON = data_dir / "current_template.json"
                web_state.ensure_data_dirs()
                web_app.CURRENT_TEMPLATE.write_bytes(_template_bytes().getvalue())
                web_state.save_template_items(read_template(web_app.CURRENT_TEMPLATE), source_name="template.xlsx")

                app = web_app.create_app()
                client = app.test_client()
                response = client.post(
                    "/api/process",
                    data={
                        "template": (_order_bytes(), "order_misfiled_as_template.xlsx"),
                        "orders": (_order_bytes(), "order.xlsx"),
                        "skipKeywords": "",
                        "fuzzyThreshold": "0.88",
                    },
                    content_type="multipart/form-data",
                )
            finally:
                web_app.RUNS_DIR = old_web_app_runs_dir
                web_app.CURRENT_TEMPLATE = old_web_app_current_template
                web_state.DATA_DIR = old_state_data_dir
                web_state.RUNS_DIR = old_state_runs_dir
                web_state.CURRENT_TEMPLATE = old_state_current_template
                web_state.TEMPLATE_JSON = old_state_template_json

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["summary"]["purchaseRows"], 1)
        self.assertEqual(payload["summary"]["unmatchedRows"], 0)
        self.assertIn("采购单.xlsx", payload["downloads"]["purchase"])
        self.assertIn("无法作为采购总模板读取", payload["warnings"][0])


def _purchase_export_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "采购单导出"
    ws.append(["采购类型", "采购负责人", "采购单号", "交货日期", "商品名称", "单位", "计划采购量", "备注"])
    ws.append(["供应商/供应商A", "客户甲", "CG1", "明天", "鸡肉", "斤", 1, None])
    ws.append(["供应商/供应商A", "客户乙", "CG1", "明天", "牛肉", "斤", 2, None])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _template_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "采购导入模板"
    ws.append(["采购导入模板"])
    ws.append(["仓库:", "默认库房"])
    ws.append(["单据备注"])
    ws.append(["负责人", "采购员/供应商名称", "商品名称", "商品单位", "采购数量", "采购单价", "商品备注"])
    ws.append(["营养餐", "供应商A", "牛肉", "斤", 0, 35.4, ""])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _order_bytes() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细导出"
    ws.append(["订单号", "客户名称", "商品名称", "发货数量", "发货单位", "订单备注"])
    ws.append(["DD1", "营养餐", "牛肉", 2, "斤", "明早送"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


if __name__ == "__main__":
    unittest.main()
