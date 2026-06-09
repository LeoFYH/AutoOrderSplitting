from __future__ import annotations

import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from auto_order_splitting.excel_io import read_orders, read_template, write_debug_workbook, write_purchase_import
from auto_order_splitting.models import PurchaseLine


class ExcelIoTests(unittest.TestCase):
    def test_reads_template_and_orders_then_writes_purchase_import_in_template_format(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            template = tmp_path / "template.xlsx"
            orders = tmp_path / "orders.xlsx"
            output = tmp_path / "out.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["采购导入模板"])
            ws.append(["仓库：", "默认库房"])
            ws.append(["单据备注"])
            ws.append(["负责人", "采购员/供应商名称", "*商品名称", "*商品单位", "*采购数量", "*采购单价", "商品备注", "已设置采购协议价"])
            ws.append(["营养餐", "供应商A", "紫菜", "包", 7, 5.5, "上午到", None])
            ws.append(["旧学校", "旧供应商", "旧商品", "斤", 999, 1, "旧备注", None])
            ws.append(["旧学校2", "旧供应商2", "旧商品2", "斤", 888, 1, "旧备注2", None])
            ws.merge_cells("A1:I1")
            ws.freeze_panes = "A20"
            ws.auto_filter.ref = "A4:I5"
            ws.row_dimensions[1].height = 25
            ws.row_dimensions[4].height = 25
            ws.column_dimensions["I"].width = 9
            wb.save(template)

            wb = Workbook()
            ws = wb.active
            ws.append(["订单号", "客户名称", "发货日期", "商品名称", "发货数量", "实际金额", "发货小计", "发货单位", "发货单价", "商品备注", "订单备注"])
            ws.append(["DD1", "学校A", "2026-06-08", "紫菜", 2, 11, 11, "包", 5.5, "不要辣", "下午到"])
            ws.append(["小计:", None, None, None, None, 11, 11, None, None, None])
            wb.save(orders)

            template_items = read_template(template)
            order_lines = read_orders([orders])
            write_purchase_import(
                template,
                output,
                [
                    PurchaseLine(
                        "学校A",
                        "供应商A",
                        "紫菜",
                        "包",
                        Decimal("2"),
                        5.5,
                        "商品备注：燕麦10件，纯牛奶10件；订单备注：号上午8：30-10点前送恒利仓库",
                        product_note="不要辣",
                        order_note="下午到",
                    )
                ],
            )

            self.assertEqual(len(template_items), 3)
            self.assertEqual(len(order_lines), 1)
            self.assertEqual(order_lines[0].product, "紫菜")
            self.assertEqual(order_lines[0].product_note, "不要辣")
            self.assertEqual(order_lines[0].order_note, "下午到")
            self.assertEqual(order_lines[0].note, "不要辣；下午到")

            written = load_workbook(output, data_only=True).active
            self.assertEqual(written.max_column, 9)
            self.assertEqual(written.max_row, 5)
            self.assertIn("A1:I1", [str(item) for item in written.merged_cells.ranges])
            self.assertEqual(written.freeze_panes, "A5")
            self.assertEqual(written.row_dimensions[1].height, 25)
            self.assertEqual(written.row_dimensions[4].height, 25)
            self.assertEqual(written.column_dimensions["I"].width, 9)
            self.assertEqual(written.cell(4, 7).value, "商品备注")
            self.assertEqual(written.cell(4, 8).value, "已设置采购协议价")
            self.assertEqual(written.cell(5, 1).value, "学校A")
            self.assertEqual(written.cell(5, 5).value, 2)
            self.assertEqual(written.cell(5, 7).value, "燕麦10件，纯牛奶10件；号上午8：30-10点前送恒利仓库；不要辣；下午到")
            self.assertIsNone(written.cell(6, 1).value)
            self.assertEqual(written.auto_filter.ref, "A4:I5")

    def test_writes_debug_workbook_with_four_tabs(self):
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "debug.xlsx"
            write_debug_workbook(
                output,
                [PurchaseLine("学校A", "供应商A", "紫菜", "包", Decimal("2"), 5.5, "下午到")],
                [],
                [],
                ["单价不同"],
            )

            wb = load_workbook(output, data_only=True)
            self.assertEqual(wb.sheetnames, ["采购结果", "未匹配", "警告", "跳过"])
            self.assertEqual(wb["采购结果"].cell(2, 1).value, "供应商A")
            self.assertEqual(wb["警告"].cell(2, 1).value, "单价不同")

    def test_reads_order_export_with_order_quantity_aliases_from_non_active_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            orders = Path(tmp) / "orders.xlsx"

            wb = Workbook()
            pivot = wb.active
            pivot.title = "透视"
            pivot.append(["求和项:下单数量", "列标签"])
            detail = wb.create_sheet("订单明细导出")
            detail.append(["订单号", "客户名称", "订单时间", "发货日期", "商品名称", "下单数量", "下单单位", "商品备注", "下单单价", "订单备注", "默认供应商"])
            detail.append(["DD1", "学校A", "2026-06-06", "2026-06-08", "紫菜", 3, "包", "不要辣", 5.5, "下午到", "订单供应商"])
            wb.save(orders)

            order_lines = read_orders([orders])

            self.assertEqual(len(order_lines), 1)
            self.assertEqual(order_lines[0].quantity, Decimal("3"))
            self.assertEqual(order_lines[0].unit, "包")
            self.assertEqual(order_lines[0].order_price, 5.5)
            self.assertEqual(order_lines[0].supplier, "订单供应商")
            self.assertEqual(order_lines[0].product_note, "不要辣")
            self.assertEqual(order_lines[0].order_note, "下午到")
            self.assertEqual(order_lines[0].note, "不要辣；下午到")

    def test_order_output_note_only_removes_labels(self):
        with tempfile.TemporaryDirectory() as tmp:
            orders = Path(tmp) / "orders.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["订单号", "客户名称", "商品名称", "下单数量", "下单单位", "商品备注", "订单备注"])
            ws.append(["DD1", "学校A", "紫菜", 3, "包", "紫菜3包；不要辣", "订单备注：下午4点送到"])
            ws.append(["DD2", "学校A", "海带", 2, "斤", "商品备注：上午送到；海带2斤", ""])
            ws.append(["DD3", "学校A", "豆腐", 5, "斤", "豆腐5斤", ""])
            ws.append(["DD4", "学校A", "牛奶", 10, "件", "商品备注：燕麦10件，纯牛奶10件；订单备注：号上午8：30-10点前送恒利仓库", ""])
            wb.save(orders)

            order_lines = read_orders([orders])

            self.assertEqual(order_lines[0].note, "紫菜3包；不要辣；下午4点送到")
            self.assertEqual(order_lines[1].note, "上午送到；海带2斤")
            self.assertEqual(order_lines[2].note, "豆腐5斤")
            self.assertEqual(order_lines[3].note, "燕麦10件，纯牛奶10件；号上午8：30-10点前送恒利仓库")


if __name__ == "__main__":
    unittest.main()
